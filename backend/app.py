from __future__ import annotations

import json
import re
import threading
from functools import lru_cache
from pathlib import Path
from subprocess import CalledProcessError
from typing import Dict, Iterable, List, Literal, Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from starlette.responses import JSONResponse
from starlette.status import HTTP_404_NOT_FOUND

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - handled during runtime only
    raise RuntimeError("openpyxl is required to run the backend") from exc

from allthatstax.config import load_config
from allthatstax.latex_text import generate_latex_text
from get_cards_information import get_cards_information
from localization import localization
from run_latex import DEFAULT_COMMAND, run_latex

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = BASE_DIR / "config.json"

LEGALITY_KEYS: List[str] = [
    "standard",
    "alchemy",
    "pioneer",
    "explorer",
    "modern",
    "historic",
    "legacy",
    "pauper",
    "vintage",
    "timeless",
    "commander",
    "duel",
]

CARD_TYPE_ORDER = ["生物", "神器", "结界", "其他"]

_mana_pattern = re.compile(r"\{([^}]+)\}")
_cache_lock = threading.Lock()
_cached_payload: Optional[Dict[str, object]] = None
_cached_mtime: Optional[float] = None


def _load_config() -> Dict[str, object]:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Config file not found at {CONFIG_PATH}")
    with CONFIG_PATH.open("r", encoding="utf-8") as fh:
        return json.load(fh)


CONFIG = _load_config()


class CardFace(BaseModel):
    englishName: str
    chineseName: str
    image: str
    manaCost: List[str]
    cardType: str
    description: str


class StaxType(BaseModel):
    key: str
    label: str


class Card(BaseModel):
    id: str
    kind: Literal["single", "multiface"]
    faces: List[CardFace]
    staxType: Optional[StaxType]
    isRestricted: bool
    legalities: Dict[str, str]
    manaValue: int
    sortCardType: str


class Metadata(BaseModel):
    staxTypes: List[StaxType]
    cardTypeOrder: List[str]


class LatexSettings(BaseModel):
    sheetFileName: str
    sheetName: str
    multifaceSheetName: str
    latexTextName: str
    latexFileName: str
    latexCommand: List[str]


class LatexGenerationRequest(BaseModel):
    sheetFileName: str
    sheetName: str
    multifaceSheetName: str
    latexTextName: str
    latexFileName: str
    latexCommand: Optional[List[str]] = None
    fetchCards: bool = False
    fetchFromScratch: bool = False
    localize: bool = False
    skipCompile: bool = False


class LatexGenerationResponse(BaseModel):
    latexTextPath: str
    pdfPath: Optional[str]
    command: List[str]
    stdout: Optional[str]
    stderr: Optional[str]


app = FastAPI(title="AllThatStax API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

images_dir = BASE_DIR / str(CONFIG["image_folder_name"])
symbols_dir = BASE_DIR / "Symbols"

if images_dir.exists():
    app.mount("/images", StaticFiles(directory=images_dir), name="images")

if symbols_dir.exists():
    app.mount("/symbols", StaticFiles(directory=symbols_dir), name="symbols")


def _normalise_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "是", "真"}


def _parse_mana_cost(raw_cost: Optional[str]) -> List[str]:
    if raw_cost is None:
        return []
    if "{" not in raw_cost:
        text = raw_cost.strip()
        return [text] if text else []
    return [token.upper() for token in _mana_pattern.findall(raw_cost)]


def _build_face(prefix: str, row: List[object]) -> CardFace:
    english = str(row[0]).strip()
    chinese = str(row[1]).strip()
    image_value = row[2] if len(row) > 2 else None
    image_name = str(image_value).strip() if image_value else ""
    image_path = f"/images/{image_name}" if image_name else ""
    mana_cost_raw = row[3] if len(row) > 3 else None
    card_type = str(row[4]).strip()
    description_raw = row[5] if len(row) > 5 else ""
    description = str(description_raw or "").strip()
    return CardFace(
        englishName=english,
        chineseName=chinese,
        image=image_path,
        manaCost=_parse_mana_cost(mana_cost_raw if isinstance(mana_cost_raw, str) else str(mana_cost_raw) if mana_cost_raw is not None else None),
        cardType=card_type,
        description=description,
    )


def _read_sheet_rows(sheet) -> Iterable[List[object]]:
    for row in sheet.iter_rows(values_only=True):
        yield list(row)


def _load_cards_payload(force: bool = False) -> Dict[str, object]:
    global _cached_payload, _cached_mtime

    sheet_path = BASE_DIR / str(CONFIG["sheet_file_name"])
    if not sheet_path.exists():
        raise FileNotFoundError(f"Sheet file not found at {sheet_path}")

    mtime = sheet_path.stat().st_mtime
    with _cache_lock:
        if not force and _cached_payload is not None and _cached_mtime == mtime:
            return _cached_payload

        workbook = load_workbook(sheet_path, data_only=True)

        single_sheet = workbook[str(CONFIG["sheet_name"])]
        multi_sheet = workbook[str(CONFIG["multiface_sheet_name"])]

        cards: List[Card] = []

        single_rows = list(_read_sheet_rows(single_sheet))
        multi_rows = list(_read_sheet_rows(multi_sheet))

        if single_rows:
            single_rows = single_rows[1:]  # remove header
        if multi_rows:
            multi_rows = multi_rows[1:]

        for row in single_rows:
            if not any(row):
                continue
            face = _build_face("", row)
            legality_values = {key: str(row[idx]).strip() if row[idx] is not None else "unknown" for idx, key in enumerate(LEGALITY_KEYS, start=8)}
            card = Card(
                id=f"single-{face.englishName}",
                kind="single",
                faces=[face],
                staxType=_build_stax_type(row[6]),
                isRestricted=_normalise_bool(row[7]),
                legalities=legality_values,
                manaValue=int(row[20]) if row[20] is not None else 0,
                sortCardType=str(row[21] or "其他"),
            )
            cards.append(card)

        for row in multi_rows:
            if not any(row):
                continue
            front = _build_face("front_", row[0:6])
            back = _build_face("back_", row[6:12])
            legality_values = {key: str(row[idx]).strip() if row[idx] is not None else "unknown" for idx, key in enumerate(LEGALITY_KEYS, start=14)}
            card = Card(
                id=f"multiface-{front.englishName}",
                kind="multiface",
                faces=[front, back],
                staxType=_build_stax_type(row[12]),
                isRestricted=_normalise_bool(row[13]),
                legalities=legality_values,
                manaValue=int(row[26]) if row[26] is not None else 0,
                sortCardType=str(row[27] or "其他"),
            )
            cards.append(card)

        stax_types = _build_stax_types()

        payload = {
            "cards": cards,
            "metadata": {
                "staxTypes": stax_types,
                "cardTypeOrder": CARD_TYPE_ORDER,
            },
        }

        _cached_payload = payload
        _cached_mtime = mtime
        return payload


def _build_stax_type(value: object) -> Optional[StaxType]:
    if value is None:
        return None
    english_key = str(value).strip()
    if not english_key:
        return None
    chinese = str(CONFIG.get("stax_type", {}).get(english_key, english_key))
    return StaxType(key=english_key, label=chinese)


@lru_cache()
def _build_stax_types() -> List[StaxType]:
    mapping = CONFIG.get("stax_type", {})
    stax_types = [StaxType(key=key, label=str(label)) for key, label in mapping.items()]
    stax_types.sort(key=lambda item: item.label)
    return stax_types


def _resolve_path_within_base(path_value: str | Path) -> Path:
    candidate = Path(path_value)
    if not candidate.is_absolute():
        candidate = BASE_DIR / candidate
    candidate = candidate.resolve()
    try:
        candidate.relative_to(BASE_DIR)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="提供的路径不在项目目录内") from exc
    return candidate


def _relative_to_base(path_value: Path) -> str:
    return str(path_value.relative_to(BASE_DIR))


@app.get("/health")
def health_check() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/cards", response_model=List[Card])
def list_cards(force_reload: bool = Query(False, alias="reload")) -> List[Card]:
    payload = _load_cards_payload(force=force_reload)
    return payload["cards"]  # type: ignore[return-value]


@app.get("/metadata", response_model=Metadata)
def get_metadata() -> Metadata:
    payload = _load_cards_payload()
    metadata = payload["metadata"]
    return Metadata(**metadata)  # type: ignore[arg-type]


@app.get("/cards/{card_id}", response_model=Card)
def get_card(card_id: str) -> Card:
    payload = _load_cards_payload()
    cards: Iterable[Card] = payload["cards"]  # type: ignore[assignment]
    for card in cards:
        if card.id == card_id:
            return card
    raise HTTPException(status_code=HTTP_404_NOT_FOUND, detail="Card not found")


@app.get("/latex/settings", response_model=LatexSettings)
def get_latex_settings() -> LatexSettings:
    config = load_config(CONFIG_PATH)
    return LatexSettings(
        sheetFileName=str(config.get("sheet_file_name", "card_information_sheet.xlsx")),
        sheetName=str(config.get("sheet_name", "Sheet")),
        multifaceSheetName=str(config.get("multiface_sheet_name", "Multiface Sheet")),
        latexTextName=str(config.get("latex_text_name", "latex_text.txt")),
        latexFileName=str(config.get("latex_file_name", "AllThatStax.tex")),
        latexCommand=list(DEFAULT_COMMAND),
    )


@app.post("/latex/generate", response_model=LatexGenerationResponse)
def generate_latex(payload: LatexGenerationRequest) -> LatexGenerationResponse:
    config = load_config(CONFIG_PATH)

    sheet_path = _resolve_path_within_base(payload.sheetFileName)
    latex_text_path = _resolve_path_within_base(payload.latexTextName)
    latex_file_path = _resolve_path_within_base(payload.latexFileName)

    card_list_path = _resolve_path_within_base(str(config.get("card_list_name", "card_list.txt")))
    image_folder_path = _resolve_path_within_base(str(config.get("image_folder_name", "Images")))

    command = payload.latexCommand[:] if payload.latexCommand else list(DEFAULT_COMMAND)

    try:
        if payload.fetchCards or payload.fetchFromScratch:
            get_cards_information(
                str(image_folder_path),
                str(sheet_path),
                payload.sheetName,
                payload.multifaceSheetName,
                str(card_list_path),
                dict(config.get("stax_type", {})),
                from_scratch=payload.fetchFromScratch,
            )

        if payload.localize:
            localization(str(sheet_path), payload.sheetName, payload.multifaceSheetName)

        latex_text_result = generate_latex_text(
            sheet_file_name=str(sheet_path),
            sheet_name=payload.sheetName,
            multiface_sheet_name=payload.multifaceSheetName,
            latex_text_name=str(latex_text_path),
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    stdout = None
    stderr = None
    pdf_path: Optional[Path] = None

    if not payload.skipCompile:
        try:
            result = run_latex(
                latex_file_name=str(latex_file_path),
                latex_text_name=str(latex_text_result),
                command=command,
            )
        except FileNotFoundError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc
        except CalledProcessError as exc:
            stdout = exc.stdout or None
            stderr = exc.stderr or None
            pdf_candidate = latex_file_path.with_suffix(".pdf")
            if pdf_candidate.exists():
                pdf_path = pdf_candidate
            return LatexGenerationResponse(
                latexTextPath=_relative_to_base(latex_text_path),
                pdfPath=_relative_to_base(pdf_path) if pdf_path else None,
                command=command,
                stdout=stdout,
                stderr=stderr,
            )
        stdout = result.stdout or None
        stderr = result.stderr or None
        pdf_candidate = latex_file_path.with_suffix(".pdf")
        if pdf_candidate.exists():
            pdf_path = pdf_candidate

    response = LatexGenerationResponse(
        latexTextPath=_relative_to_base(latex_text_path),
        pdfPath=_relative_to_base(pdf_path) if pdf_path else None,
        command=command,
        stdout=stdout,
        stderr=stderr,
    )

    return response


@app.get("/latex/download")
def download_latex_pdf(path: str = Query(..., description="相对于项目根目录的 PDF 路径")) -> FileResponse:
    file_path = _resolve_path_within_base(path)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="指定的文件不存在")
    if file_path.suffix.lower() != ".pdf":
        raise HTTPException(status_code=400, detail="仅支持下载 PDF 文件")
    return FileResponse(file_path, filename=file_path.name, media_type="application/pdf")


@app.exception_handler(FileNotFoundError)
async def _handle_missing_file(_: "FileNotFoundError") -> JSONResponse:
    return JSONResponse(status_code=500, content={"detail": "Required data file is missing."})
