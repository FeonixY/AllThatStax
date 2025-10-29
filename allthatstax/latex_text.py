"""Helpers for turning workbook data into LaTeX content."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["generate_latex_text"]

CARD_TYPE_ORDER: Dict[str, int] = {
    "生物": 1,
    "神器": 2,
    "结界": 3,
    "其他": 4,
}

LEGALITY_FIELDS: Sequence[str] = (
    "standard",
    "alchemy",
    "pioneer",
    "explorer",
    "modern",
    "historic",
    "legacy",
    "pauper",
    "vintage",
    "timeless",
    "commander",
    "duel_commander",
)


@dataclass
class LatexCard:
    body: str
    mana_value: int
    sort_type: str
    english_name: str

    def base_sort_key(self) -> tuple[int, int, str]:
        return (
            self.mana_value,
            CARD_TYPE_ORDER.get(self.sort_type, len(CARD_TYPE_ORDER) + 1),
            self.english_name,
        )

    def plus_sort_key(self) -> tuple[int, int, str]:
        return (
            CARD_TYPE_ORDER.get(self.sort_type, len(CARD_TYPE_ORDER) + 1),
            self.mana_value,
            self.english_name,
        )


def _iter_values(sheet: Worksheet) -> Iterator[Sequence[object]]:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            yield row


def _coerce_text(value: object, default: str = "") -> str:
    if value is None:
        return default
    text = str(value)
    if text.strip().lower() == "none":
        return default
    return text


def _format_mana_cost(value: object) -> str:
    text = _coerce_text(value)
    if not text:
        return "无费用（法术力值为0）"
    return text.replace("{", "\\MTGsymbol{").replace("}", "}{5}")


def _format_description(value: object) -> str:
    text = _coerce_text(value)
    if not text:
        return ""
    return text.replace("{", "\\MTGsymbol{").replace("}", "}{3}").replace("\n", "\\\\\n")


def _get_value(values: Sequence[object], index: int, default: object = "") -> object:
    return values[index] if index < len(values) else default


def _parse_mana_value(value: object) -> int:
    text = _coerce_text(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def _format_legalities(values: Sequence[object]) -> str:
    rendered: List[str] = []
    for index, label in enumerate(LEGALITY_FIELDS):
        entry = _get_value(values, index)
        suffix = "," if index < len(LEGALITY_FIELDS) - 1 else ""
        rendered.append(f"\tlegality / {label} = {_coerce_text(entry)}{suffix}")
    return "\n".join(rendered)


def _build_single_cards(sheet: Worksheet) -> Iterable[LatexCard]:
    for row in _iter_values(sheet):
        row_list = list(row)
        english_name = _coerce_text(_get_value(row_list, 0))
        if not english_name:
            continue
        chinese_name = _coerce_text(_get_value(row_list, 1))
        image_path = _coerce_text(_get_value(row_list, 2))
        mana_cost = _format_mana_cost(_get_value(row_list, 3))
        card_type = _coerce_text(_get_value(row_list, 4))
        description = _format_description(_get_value(row_list, 5))
        stax_type = _coerce_text(_get_value(row_list, 6))
        restricted = _coerce_text(_get_value(row_list, 7))
        legalities = [
            _get_value(row_list, index)
            for index in range(8, 8 + len(LEGALITY_FIELDS))
        ]
        mana_value = _parse_mana_value(_get_value(row_list, 20))
        sort_type = _coerce_text(_get_value(row_list, 21), "其他")

        lines = [
            "\\card",
            "{",
            f"\tcard_english_name = {{{english_name}}},",
            f"\tcard_chinese_name = {{{chinese_name}}},",
            f"\tcard_image = {image_path},",
            f"\tmana_cost = {mana_cost},",
            f"\tcard_type = {card_type},",
            f"\tdescription = {{{description}}},",
            f"\tstax_type = {stax_type},",
            f"\tis_in_restricted_list = {restricted},",
            _format_legalities(legalities),
            "}",
            "",
        ]
        yield LatexCard("\n".join(lines), mana_value, sort_type, english_name)


def _build_multiface_cards(sheet: Worksheet) -> Iterable[LatexCard]:
    for row in _iter_values(sheet):
        row_list = list(row)
        english_front = _coerce_text(_get_value(row_list, 0))
        if not english_front:
            continue
        chinese_front = _coerce_text(_get_value(row_list, 1))
        image_front = _coerce_text(_get_value(row_list, 2))
        mana_front = _format_mana_cost(_get_value(row_list, 3))
        type_front = _coerce_text(_get_value(row_list, 4))
        desc_front = _format_description(_get_value(row_list, 5))
        english_back = _coerce_text(_get_value(row_list, 6))
        chinese_back = _coerce_text(_get_value(row_list, 7))
        image_back = _coerce_text(_get_value(row_list, 8))
        mana_back = _format_mana_cost(_get_value(row_list, 9))
        type_back = _coerce_text(_get_value(row_list, 10))
        desc_back = _format_description(_get_value(row_list, 11))
        stax_type = _coerce_text(_get_value(row_list, 12))
        restricted = _coerce_text(_get_value(row_list, 13))
        legalities = [
            _get_value(row_list, index)
            for index in range(14, 14 + len(LEGALITY_FIELDS))
        ]
        mana_value = _parse_mana_value(_get_value(row_list, 26))
        sort_type = _coerce_text(_get_value(row_list, 27), "其他")

        lines = [
            "\\mfcard",
            "{",
            f"\tfront_card_english_name = {{{english_front}}},",
            f"\tfront_card_chinese_name = {{{chinese_front}}},",
            f"\tfront_card_image = {image_front},",
            f"\tfront_mana_cost = {mana_front},",
            f"\tfront_card_type = {type_front},",
            f"\tfront_description = {{{desc_front}}},",
            f"\tback_card_english_name = {{{english_back}}},",
            f"\tback_card_chinese_name = {{{chinese_back}}},",
            f"\tback_card_image = {image_back},",
            f"\tback_mana_cost = {mana_back},",
            f"\tback_card_type = {type_back},",
            f"\tback_description = {{{desc_back}}},",
            f"\tstax_type = {stax_type},",
            f"\tis_in_restricted_list = {restricted},",
            _format_legalities(legalities),
            "}",
            "",
        ]
        yield LatexCard("\n".join(lines), mana_value, sort_type, english_front)


def generate_latex_text(
    sheet_file_name: str | Path,
    sheet_name: str,
    multiface_sheet_name: str,
    latex_text_name: str | Path,
) -> Path:
    """Generate LaTeX snippets from the workbook and write them to disk."""

    workbook_path = Path(sheet_file_name)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    try:
        singles_sheet = workbook[sheet_name]
    except KeyError as exc:  # pragma: no cover - invalid configuration
        raise ValueError(f"Worksheet '{sheet_name}' not found in {workbook_path}") from exc
    try:
        multiface_sheet = workbook[multiface_sheet_name]
    except KeyError as exc:  # pragma: no cover - invalid configuration
        raise ValueError(f"Worksheet '{multiface_sheet_name}' not found in {workbook_path}") from exc

    cards: List[LatexCard] = []
    cards.extend(_build_single_cards(singles_sheet))
    cards.extend(_build_multiface_cards(multiface_sheet))
    workbook.close()

    cards.sort(key=lambda card: card.base_sort_key())

    groups: Dict[object, List[LatexCard]] = {i: [] for i in range(7)}
    groups["7+"] = []

    for card in cards:
        if card.mana_value >= 7:
            groups["7+"].append(card)
        else:
            groups.setdefault(card.mana_value, []).append(card)

    groups["7+"].sort(key=lambda card: card.plus_sort_key())

    output_path = Path(latex_text_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as handle:
        for cmc in range(1, 7):
            handle.write(f"\\chapter{{{cmc}费}}\n\n")
            current_type = None
            for card in groups.get(cmc, []):
                if card.sort_type != current_type:
                    current_type = card.sort_type
                    handle.write(f"\\section{{{current_type}}}\n\n")
                handle.write(card.body)
        handle.write("\\chapter{7+费}\n\n")
        current_type = None
        for card in groups["7+"]:
            if card.sort_type != current_type:
                current_type = card.sort_type
                handle.write(f"\\section{{{current_type}}}\n\n")
            handle.write(card.body)
        handle.write("\\chapter{0费（包括地）}\n\n")
        current_type = None
        for card in groups.get(0, []):
            if card.sort_type != current_type:
                current_type = card.sort_type
                handle.write(f"\\section{{{current_type}}}\n\n")
            handle.write(card.body)

    return output_path
