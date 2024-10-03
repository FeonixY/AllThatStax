import re
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def localization(
    sheet_file_name: str,
    sheet_name: str,
    multiface_sheet_name: str):

    workbook = load_workbook(sheet_file_name)
    
    sheet = workbook[sheet_name]
    multiface_sheet = workbook[multiface_sheet_name]

    serial_number_pattern = re.compile(r'([a-zA-Z0-9]+)-(\d+)-')
    
    driver = webdriver.Chrome(service = Service(ChromeDriverManager().install()))

    for row in sheet.iter_rows(min_row = 2):
        card_chinese_name = row[1].value
        card_image_name  = row[2].value
        card_type = row[4].value
        card_description = row[5].value
        
        if not card_chinese_name:
            try:
                match = serial_number_pattern.search(card_image_name)
                if match:
                    series_code = match.group(1).upper()
                    card_number = match.group(2)
                    card_serial_number = f"{series_code}/{card_number}"
                else:
                    print(f"Invalid card image name format: {card_image_name}")
                    continue
                
                url = f"https://sbwsz.com/card/{card_serial_number}"
                driver.get(url)

                card_section = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "section.el-container.container"))
                )

                card_chinese_name = card_section.find_element(By.CSS_SELECTOR, 'main.el-main.card-name.name-zh > span.el-text').text
                card_type = card_section.find_element(By.CSS_SELECTOR, 'div.card-type > span.el-text').text
                card_description = card_section.find_element(By.CSS_SELECTOR, 'div.card-text-container > span.el-text').text

                row[1].value = card_chinese_name
                row[4].value = card_type
                row[5].value = card_description

                print(f"Updated card info for {card_serial_number}: {card_chinese_name}, {card_type}, {card_description}")

            except Exception as e:
                print(f"Error retrieving card info for {card_serial_number}: {e}")
    
    for row in multiface_sheet.iter_rows(min_row = 2):
        front_card_chinese_name = row[1].value
        front_card_image_name  = row[2].value
        front_card_type = row[4].value
        front_card_description = row[5].value
        back_card_chinese_name = row[7].value
        back_card_type = row[10].value
        back_card_description = row[11].value
        
        if not front_card_chinese_name:
            try:
                match = serial_number_pattern.search(front_card_image_name)
                if match:
                    series_code = match.group(1).upper()
                    card_number = match.group(2)
                    card_serial_number = f"{series_code}/{card_number}"
                else:
                    print(f"Invalid card image name format: {front_card_image_name}")
                    continue
                
                url = f"https://sbwsz.com/card/{card_serial_number}"
                driver.get(url)

                card_sections = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.el-container.container")))

                try:
                    front_card_chinese_name = card_sections[0].find_element(By.CSS_SELECTOR, 'main.el-main.card-name.name-zh > span.el-text').text
                    front_card_type = card_sections[0].find_element(By.CSS_SELECTOR, 'div.card-type > span.el-text').text
                    front_card_description = card_sections[0].find_element(By.CSS_SELECTOR, 'div.card-text-container > span.el-text').text
                    back_card_chinese_name = card_sections[1].find_element(By.CSS_SELECTOR, 'main.el-main.card-name.name-zh > span.el-text').text
                    back_card_type = card_sections[1].find_element(By.CSS_SELECTOR, 'div.card-type > span.el-text').text
                    back_card_description = card_sections[1].find_element(By.CSS_SELECTOR, 'div.card-text-container > span.el-text').text

                    row[1].value = front_card_chinese_name
                    row[4].value = front_card_type
                    row[5].value = front_card_description
                    row[7].value = back_card_chinese_name
                    row[10].value = back_card_type
                    row[11].value = back_card_description

                    print(f"Updated card info for {card_serial_number}: " \
                          f"{front_card_chinese_name}, {front_card_type}, {front_card_description}, " \
                          f"{back_card_chinese_name}, {back_card_type}, {back_card_description}")

                except Exception as e:
                    print(f"Error extracting card information: {e}")
                    
            except Exception as e:
                print(f"Error retrieving card info for {card_serial_number}: {e}")

    workbook.save(sheet_file_name)

    driver.quit()

