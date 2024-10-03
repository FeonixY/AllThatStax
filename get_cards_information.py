import os
import re
import time
import requests
from openpyxl import load_workbook
from urllib.parse import urlparse, unquote

def get_cards_information(
    image_folder_name: str,
    sheet_file_name : str,
    sheet_name : str,
    multiface_sheet_name : str,
    card_list_name : str,
    stax_type_dict : dict,
    from_scratch : bool = False):
    
    def get_card_json_data(card_name):
        # 整理卡牌名称
        card_name_match = re.search(r'\d+\s+(.+?)\s+\(.*?\)', card_name)
        if card_name_match:
            card_english_name = card_name_match.group(1)
        tags = re.findall(r'#(\S.+?)\s*(?=#|$)', card_name)

        # 检索中文信息
        url = f"https://api.scryfall.com/cards/search?q=lang:zhs%20!%22{card_english_name}%22"
        response = requests.get(url)

        # 如果存在中文信息，则使用
        if response.status_code == 200:
            return response.json()["data"][0], tags

        # 如果不存在中文信息，则检索并使用默认信息
        elif response.status_code == 404:
            new_url = f"https://api.scryfall.com/cards/named?exact={card_english_name}"
            new_response = requests.get(new_url)
            if new_response.status_code != 200:
                print(f"Failed to get {card_english_name} Chinese card name or other versions. {new_response.status_code} - {new_response.text}")
                return None
            return new_response.json(), tags

        # 如果也不存在，则获取卡牌信息失败，跳过这张卡牌
        else:
            print(f"Failed to get {card_english_name} information. {response.status_code} - {response.text}")
            return None
    
    def get_card_image(image_uri, card_english_name):
        # 获取图片名
        image_name_response = requests.head(image_uri, allow_redirects = True)
        if image_name_response.status_code == 200:
            if "Content-Disposition" in image_name_response.headers:
                content_disposition = image_name_response.headers["Content-Disposition"]
                card_image_name = content_disposition.split("filename=")[1].strip("'").strip('"')
            else:
                parsed_url = urlparse(image_uri)
                card_image_name = os.path.basename(unquote(parsed_url.path))
                print(f"{card_english_name} URL doesn't contain Content-Disposition, using name parsed from URL as filename")
                        
            # 获取图片
            os.makedirs(image_folder_name, exist_ok = True)
            image_response = requests.get(image_uri)
            if image_response.status_code == 200:
                with open(f"{image_folder_name}/{card_image_name}", "wb") as f:
                    f.write(image_response.content)
            else:
                card_image_name = None
                print(f"Failed to download {card_english_name} image. {image_response.status_code} - {image_response.text}")
        else:
            card_image_name = None
            print(f"Failed to get {card_english_name} image headers. {image_name_response.status_code} - {image_name_response.text}")

        used_images.add(card_image_name)
        return card_image_name

    type_dict = {
        "Creature": "生物",
        "Artifact": "神器",
        "Enchantment": "结界"
    }

    def process_card_data(card_data, tags):
        if card_data is None:
            return None
        
        def replace_not_legal_in_legalities(original_legalities):
            def replace_values(obj):
                if isinstance(obj, dict):
                    return {k: replace_values(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [replace_values(item) for item in obj]
                elif obj == "not_legal":
                    return "notlegal"
                else:
                    return obj

            return replace_values(original_legalities)
        
        legalities = replace_not_legal_in_legalities(card_data["legalities"])
        if "card_faces" in card_data:
            front_card_data = card_data["card_faces"][0]
            back_card_data = card_data["card_faces"][1]

            """
            双面牌表格结构：
            front_card_english_name, front_card_chinese_name, front_card_image_name, front_mana_cost, front_card_type, front_card_description,
            back_card_english_name, back_card_chinese_name, back_card_image_name, back_mana_cost, back_card_type, back_card_description,
            stax_type, is_rl, 
            legalities["standard"], legalities["alchemy"], legalities["pioneer"], legalities["explorer"],
            legalities["modern"], legalities["historic"], legalities["legacy"], legalities["pauper"],
            legalities["vintage"], legalities["timeless"], legalities["commander"], legalities["duel"],
            cmc, sort_card_type
            """

            return (True, [
                front_card_data["name"],
                front_card_data["printed_name"] if "printed_name" in front_card_data else "",
                get_card_image(front_card_data["image_uris"]["png"] if card_data["layout"] != "flip" else card_data["image_uris"]["png"], front_card_data["name"]),
                front_card_data["mana_cost"],
                front_card_data["printed_type_line"] if "printed_type_line" in front_card_data else front_card_data["type_line"],
                re.sub(r"\([^)]*\)|（[^）]*）", "", front_card_data["printed_text"]) if "printed_text" in front_card_data else re.sub(r"\([^)]*\)|（[^）]*）", "", front_card_data["oracle_text"]),
                back_card_data["name"],
                back_card_data["printed_name"] if "printed_name" in back_card_data else "",
                get_card_image(back_card_data["image_uris"]["png"] if card_data["layout"] != "flip" else card_data["image_uris"]["png"], back_card_data["name"]),
                back_card_data["mana_cost"],
                back_card_data["printed_type_line"] if "printed_type_line" in back_card_data else back_card_data["type_line"],
                re.sub(r"\([^)]*\)|（[^）]*）", "", back_card_data["printed_text"]) if "printed_text" in back_card_data else re.sub(r"\([^)]*\)|（[^）]*）", "", back_card_data["oracle_text"]),
                "，".join(stax_type_dict[tag] for tag in tags if tag in stax_type_dict),
                "RL" if card_data["reserved"] else "Not RL",
                legalities["standard"], legalities["alchemy"], legalities["pioneer"], legalities["explorer"],
                legalities["modern"], legalities["historic"], legalities["legacy"], legalities["pauper"],
                legalities["vintage"], legalities["timeless"], legalities["commander"], legalities["duel"],
                card_data["cmc"], next((type_dict[key] for key in type_dict if key in front_card_data["type_line"]), "其他")
            ])
        else:

            """
            单面牌表格结构：
            card_english_name, card_chinese_name, card_image_name, mana_cost, card_type, card_description,
            stax_type, is_rl, 
            legalities["standard"], legalities["alchemy"], legalities["pioneer"], legalities["explorer"],
            legalities["modern"], legalities["historic"], legalities["legacy"], legalities["pauper"],
            legalities["vintage"], legalities["timeless"], legalities["commander"], legalities["duel"],
            cmc, sort_card_type
            """

            return (False, [
                card_data["name"],
                card_data["printed_name"] if "printed_name" in card_data else "",
                get_card_image(card_data["image_uris"]["png"], card_data["name"]),
                card_data["mana_cost"],
                card_data["printed_type_line"] if "printed_type_line" in card_data else card_data["type_line"],
                re.sub(r"\([^)]*\)|（[^）]*）", "", card_data["printed_text"]) if "printed_text" in card_data else re.sub(r"\([^)]*\)|（[^）]*）", "", card_data["oracle_text"]),
                "，".join(stax_type_dict[tag] for tag in tags if tag in stax_type_dict),
                "RL" if card_data["reserved"] else "Not RL",
                legalities["standard"], legalities["alchemy"], legalities["pioneer"], legalities["explorer"],
                legalities["modern"], legalities["historic"], legalities["legacy"], legalities["pauper"],
                legalities["vintage"], legalities["timeless"], legalities["commander"], legalities["duel"],
                card_data["cmc"], next((type_dict[key] for key in type_dict if key in card_data["type_line"]), "其他")
            ])
    
    # 删除Images文件夹中未用到的图片
    def clean_up_images(image_folder, used_images):
        for image_file in os.listdir(image_folder):
            if image_file not in used_images:
                os.remove(os.path.join(image_folder, image_file))
                print(f"Deleted unused image: {image_file}")

    # 删除之前的表格，并将临时表格重命名为原来的表格的名字
    def rename_sheets(workbook, sheet_name, multiface_sheet_name):
        del workbook[sheet_name]
        del workbook[multiface_sheet_name]
        workbook[sheet_name + "_temp"].title = sheet_name
        workbook[multiface_sheet_name + "_temp"].title = multiface_sheet_name

    # 读取列表
    with open(card_list_name, "r", encoding = "utf-8") as f:
        card_names = f.readlines()

    # 读取表格
    workbook = load_workbook(sheet_file_name)
    if from_scratch:
        sheet = workbook[sheet_name]
        sheet.delete_rows(2, sheet.max_row - 1)
        multiface_sheet = workbook[multiface_sheet_name]
        multiface_sheet.delete_rows(2, multiface_sheet.max_row - 1)
    else:
        prev_sheet = workbook[sheet_name]
        prev_multiface_sheet = workbook[multiface_sheet_name]
        sheet = workbook.create_sheet(f"{sheet_name}_temp")
        sheet.append([cell.value for cell in prev_sheet[1]])
        multiface_sheet = workbook.create_sheet(f"{multiface_sheet_name}_temp")
        multiface_sheet.append([cell.value for cell in prev_multiface_sheet[1]])
        workbook.save(sheet_file_name)

    # 遍历列表
    used_images = set()
    for card_name in card_names:
        # 获取卡牌信息
        if from_scratch:
            card_data, tags = get_card_json_data(card_name)

            # 处理卡牌信息并保存
            new_line = process_card_data(card_data, tags)
        else:
            is_prev_has_information = False
            for cell in prev_sheet["A"]:
                if cell.value in card_name:
                    new_line = (False, [cell.value for cell in prev_sheet[cell.row]])
                    used_images.add(prev_sheet.cell(cell.row, 3))
                    is_prev_has_information = True
                    break
            for cell in prev_multiface_sheet["A"]:
                if cell.value in card_name:
                    new_line = (True, [cell.value for cell in prev_multiface_sheet[cell.row]])
                    used_images.add(prev_multiface_sheet.cell(cell.row, 3))
                    used_images.add(prev_multiface_sheet.cell(cell.row, 8))
                    is_prev_has_information = True
                    break
            if is_prev_has_information == False:
                card_data, tags = get_card_json_data(card_name)
                new_line = process_card_data(card_data, tags)

        if new_line[0]:
            multiface_sheet.append(new_line[1])
        else:
            sheet.append(new_line[1])
        workbook.save(sheet_file_name)
        
        print(f"Add {new_line[1]}")

        # 防止触发scryfall保护机制
        time.sleep(0.3)
    
    clean_up_images(image_folder_name, used_images)
    if from_scratch == False:
        # 清理未使用的图片
        
        # 删除旧表并重命名新表
        rename_sheets(workbook, sheet_name, multiface_sheet_name)

        workbook.save(sheet_file_name)

